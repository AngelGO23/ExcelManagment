from openpyxl import Workbook, load_workbook

workbook = load_workbook(filename="Auraria_kWh_Data.xlsx")


#-----------------------------------------------------------------------------------------------------------
'''I want to iterate through every sheet in the excel file.
This is what the listOfSheets function does.'''
def listOfSheets(book= Workbook):
    #The .sheetnames function returns a list with all sheets name
    sheets = book.sheetnames
    sheetObjs = []
    for sname in sheets:
        #Workbook objects work like dictionaries
        # in the sense of key, value pair
        sheetObjs.append(book[sname])

    return sheetObjs
'''Use this next for loop to see that the sheets variable holds all the sheets
objects inside the excel file

sheets = listOfSheets(workbook)
for sheet in sheets:
    print(sheet.title)'''
#-----------------------------------------------------------------------------------------------------------
'''This functions allows the code to view the values in each row.'''
def iterRows(ws):
    '''iter_rows returns an iterator of cell objects that are inside a row which
    in turn is inside a sheet (ws in this case)'''
    for row in ws.iter_rows():
        yield [cell.value for cell in row]
#-----------------------------------------------------------------------------------------------------------
'''These 3 lines were for testing how we could obtain the time
from the first cell in the row (this case, row in index 0).

time = str(rows[0][0])

print(time)

OutPut: 2020-09-10 18:00:00
'''
'''This function is for accesing the first column of any row.
We need it to be of string type.'''
def dateAndTime(theRows=list):
    for row in theRows:
        yield str(row[0])

#-----------------------------------------------------------------------------------------------------------
'''We will combine the functions we have so far in order to
obtain all the index of the intervals in the daytime.
These indexes are use to later obtain the kWh values.'''
def findIndxHour(allSheets):

    for i in range(0, len(allSheets)):
        indx4Daytime = []
        rows = list(iterRows(allSheets[i]))
        intervalTimes = list(dateAndTime(rows))
        count = 0
        for time in intervalTimes:

            count += 1
            temp2 = time.split()
            if temp2[0] != 'None':
                HrMinSec = temp2[1].split(":")
                #Cast as int so it can be used for comparison
                hour = int(HrMinSec[0])
                if hour >= 6 and hour <= 18:
                    indx4Daytime.append(count)
        yield indx4Daytime

'''This section is the proof of concept for finding the hour.
Which we later use to find the time we need for finding the
correct kWh value. These kWh values will help use identify the 
daytime minimum load.

time = str(rows[0][0])

splitted = time.split()
print(splitted)

HrMinSec = splitted[1].split(":")
print(HrMinSec)

hour = HrMinSec[0]
print(hour)
'''
#-----------------------------------------------------------------------------------------------------------

'''Use these two lines to look at each functions individual output.

# rows will store the values of each row as a list
rows = list(iterRows(sheets[0]))


# IntervalTimes will store a list of the date and the time of x interval
IntervalTimes = list(dateAndTime(rows))

# allIndex holds all the indexes of the daytime kWh values
allIndex = list(findIndxHour(sheets))
'''

#-----------------------------------------------------------------------------------------------------------
'''daytimekWh is function for accesing the value of the second column in each row.
We need it to be of int type.'''

def daytimekWh(theRows=list):
    for row in theRows:
        yield str(row[1])

#-----------------------------------------------------------------------------------------------------------
'''allDaytime returns the minimum load of all loads in the daytime intervals.'''
def allDaytimekWh(allSheets):
    #For finding the daytime indexes
    allIndex = list(findIndxHour(allSheets))
    for i in range(0, len(allSheets)):
        #Temp will hold the kWh values of the daytime intervals
        temp = []
        print(allSheets[i].title)
        rows = list(iterRows(allSheets[i]))
        #kWh holds ALL kWh values in the excel sheet
        kWh = list(daytimekWh(rows))
        for z in range(0, len(kWh)):
            # Checks if the index i is inside the first element of the allIndex list
            # Uses allIndex[0] because all sheets have the same index for their daytime.
            if allIndex[0].__contains__(z) and kWh[z] != 'None' and float(kWh[z]) != 0:
                # if it is a daytime interval, it adds it to the temp list
                temp.append(float(kWh[z]))

        # yields the minimum of the temp list for the ith sheet
        yield min(temp)

#-----------------------------------------------------------------------------------------------------------
'''Putting everything together'''

sheets = listOfSheets(workbook)

daytimeMinLoad = list(allDaytimekWh(sheets))

# result stores a dictionary with the building as Key and its daytime minimum as value
result = {}

for i in range(0, len(sheets)):
    result[sheets[i].title] = int(daytimeMinLoad[i])

print(result)


